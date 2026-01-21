from __future__ import annotations

import logging
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from .config import BASE_DIR, settings
from .db import get_hist_conn, init_historical_db
from .utils import week_start_monday


log = logging.getLogger("historical")

BRANCH_CODE_MAP = {
    "СМ": 1213086,
    "МП": 1224674,
    "СС": 1224689,
}

_LATIN_TO_CYRILLIC = str.maketrans(
    {
        "A": "А",
        "B": "В",
        "C": "С",
        "E": "Е",
        "H": "Н",
        "K": "К",
        "M": "М",
        "O": "О",
        "P": "Р",
        "T": "Т",
        "X": "Х",
        "Y": "У",
    }
)

_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$")


@dataclass
class FileInfo:
    path: str
    exists: bool
    size: int | None
    mtime: float | None


def _normalize_prefix(value: str) -> str:
    text = value.strip().upper().translate(_LATIN_TO_CYRILLIC)
    return text


def _parse_sheet_name(name: str) -> tuple[int, str] | None:
    cleaned = name.strip()
    parts = [part.strip() for part in cleaned.split(".")]
    if len(parts) != 3:
        return None
    prefix_raw, month_raw, year_raw = parts
    if not (month_raw.isdigit() and year_raw.isdigit()):
        return None
    prefix = _normalize_prefix(prefix_raw)
    branch_id = BRANCH_CODE_MAP.get(prefix)
    if not branch_id:
        return None
    month = int(month_raw)
    if not (1 <= month <= 12):
        return None
    year = int(year_raw)
    if year < 100:
        year += 2000
    month_key = f"{year:04d}-{month:02d}"
    return branch_id, month_key



def _iter_day_columns(ws) -> list[tuple[int, date]]:
    days: list[tuple[int, date]] = []
    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is None:
            if days:
                break
            continue
        if isinstance(value, datetime):
            days.append((col, value.date()))
        elif isinstance(value, date):
            days.append((col, value))
    return days


def _parse_time_label(label: str) -> int | None:
    match = _TIME_RE.match(label.strip())
    if not match:
        return None
    hour = int(match.group(1))
    return hour


def _resolve_excel_path() -> Path:
    path = settings.historical_excel_path.expanduser()
    try:
        return path.resolve()
    except Exception:
        return path


def file_info() -> FileInfo:
    path = _resolve_excel_path()
    if path.exists():
        stat = path.stat()
        return FileInfo(
            path=str(path),
            exists=True,
            size=stat.st_size,
            mtime=stat.st_mtime,
        )
    return FileInfo(path=str(path), exists=False, size=None, mtime=None)


def list_root_files() -> list[str]:
    entries = []
    try:
        for item in BASE_DIR.iterdir():
            if item.is_file():
                entries.append(item.name)
    except Exception as exc:  # noqa: BLE001
        log.warning("Failed to list root files: %s", exc)
    return sorted(entries)


def _start_import(mode: str) -> str:
    init_historical_db()
    run_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    info = file_info()
    with get_hist_conn() as conn:
        conn.execute(
            """
            INSERT INTO historical_imports(run_id, started_at, status, rows_count, file_path, file_mtime, error_log)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (run_id, now, "running", 0, info.path, info.mtime, ""),
        )
        conn.commit()
    log.info("Historical import started: %s (mode=%s)", run_id, mode)
    return run_id


def _finish_import(run_id: str, status: str, rows_count: int = 0, error: str | None = None) -> None:
    now = datetime.utcnow().isoformat()
    fields = ["finished_at = ?", "status = ?", "rows_count = ?"]
    params: list[Any] = [now, status, rows_count]
    if error:
        fields.append("error_log = COALESCE(error_log, '') || ?")
        params.append(f"\n{error}")
    params.append(run_id)
    with get_hist_conn() as conn:
        conn.execute(
            f"UPDATE historical_imports SET {', '.join(fields)} WHERE run_id = ?",
            params,
        )
        conn.commit()


def _store_type_order(conn, branch_id: int, month: str, types: list[str]) -> None:
    rows = [(branch_id, month, name, idx) for idx, name in enumerate(types)]
    if rows:
        conn.executemany(
            """
            INSERT OR REPLACE INTO historical_types
            (branch_id, month, resource_type, order_index)
            VALUES (?, ?, ?, ?)
            """,
            rows,
        )


def _iter_sheet_rows(ws, branch_id: int, month: str) -> tuple[list[str], list[tuple[Any, ...]]]:
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        return [], []
    day_columns: list[tuple[int, date]] = []
    for idx, value in enumerate(first_row, start=1):
        if idx == 1:
            continue
        if value is None:
            if day_columns:
                break
            continue
        if isinstance(value, datetime):
            day_columns.append((idx, value.date()))
        elif isinstance(value, date):
            day_columns.append((idx, value))
    if not day_columns:
        return [], []

    day_indexes = [(col_idx - 1, day_date) for col_idx, day_date in day_columns]
    type_order: list[str] = []
    current_type: str | None = None
    current_rows: list[tuple[int, tuple[Any, ...]]] = []
    rows_out: list[tuple[Any, ...]] = []

    def flush_type() -> None:
        nonlocal current_type, current_rows
        if not current_type or not current_rows:
            current_rows = []
            return
        for hour, row_vals in current_rows:
            for col_idx, day_date in day_indexes:
                try:
                    val = row_vals[col_idx] if col_idx < len(row_vals) else None
                    load_pct = float(val) if val is not None else 0.0
                except (TypeError, ValueError):
                    load_pct = 0.0
                rows_out.append(
                    (
                        branch_id,
                        month,
                        current_type,
                        day_date.isoformat(),
                        day_date.isoweekday(),
                        hour,
                        load_pct,
                    )
                )
        current_rows = []

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals:
            continue
        cell_val = row_vals[0]
        if cell_val is None:
            continue
        label = cell_val.strip() if isinstance(cell_val, str) else str(cell_val).strip()
        if not label:
            continue
        hour = _parse_time_label(label)
        if hour is None:
            flush_type()
            current_type = label
            if current_type not in type_order:
                type_order.append(current_type)
            continue
        if current_type:
            current_rows.append((hour, row_vals))
    flush_type()
    return type_order, rows_out



def run_import(run_id: str, mode: str = "replace") -> None:
    init_historical_db()
    info = file_info()
    if not info.exists:
        _finish_import(run_id, "failed", error=f"Файл не найден: {info.path}")
        return
    path = Path(info.path)
    total_rows = 0
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        with get_hist_conn() as conn:
            if mode == "replace":
                conn.execute("DELETE FROM historical_loads")
                conn.execute("DELETE FROM historical_types")
            for sheet_name in wb.sheetnames:
                parsed = _parse_sheet_name(sheet_name)
                if not parsed:
                    continue
                branch_id, month_key = parsed
                ws = wb[sheet_name]
                type_order, rows_iter = _iter_sheet_rows(ws, branch_id, month_key)
                _store_type_order(conn, branch_id, month_key, type_order)
                batch: list[tuple[Any, ...]] = []
                for row in rows_iter:
                    batch.append(row)
                    if len(batch) >= 5000:
                        conn.executemany(
                            """
                            INSERT OR REPLACE INTO historical_loads
                            (branch_id, month, resource_type, date, dow, hour, load_pct)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                            """,
                            batch,
                        )
                        total_rows += len(batch)
                        batch = []
                if batch:
                    conn.executemany(
                        """
                        INSERT OR REPLACE INTO historical_loads
                        (branch_id, month, resource_type, date, dow, hour, load_pct)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        batch,
                    )
                    total_rows += len(batch)
            conn.commit()
        _finish_import(run_id, "success", rows_count=total_rows)
    except Exception as exc:  # noqa: BLE001
        log.exception("Historical import failed: %s", exc)
        _finish_import(run_id, "failed", rows_count=total_rows, error=str(exc))


def start_import(mode: str = "replace") -> str:
    return _start_import(mode)


def last_import_status() -> dict[str, Any]:
    info = file_info()
    with get_hist_conn() as conn:
        cur = conn.execute(
            """
            SELECT run_id, started_at, finished_at, status, rows_count, file_path, file_mtime, error_log
            FROM historical_imports
            ORDER BY started_at DESC
            LIMIT 1
            """
        )
        row = cur.fetchone()
    status = dict(row) if row else None
    return {
        "file": {
            "exists": info.exists,
            "path": info.path,
            "size": info.size,
            "mtime": info.mtime,
        },
        "import": status,
    }


def list_branches() -> list[dict[str, Any]]:
    with get_hist_conn() as conn:
        cur = conn.execute("SELECT DISTINCT branch_id FROM historical_loads ORDER BY branch_id")
        branch_ids = [int(r[0]) for r in cur.fetchall()]
    branches = []
    for code, branch_id in BRANCH_CODE_MAP.items():
        if branch_id in branch_ids:
            branches.append({"branch_id": branch_id, "code": code})
    for bid in branch_ids:
        if bid not in BRANCH_CODE_MAP.values():
            branches.append({"branch_id": bid, "code": None})
    return branches


def list_months(branch_id: int) -> list[str]:
    with get_hist_conn() as conn:
        cur = conn.execute(
            "SELECT DISTINCT month FROM historical_loads WHERE branch_id = ? ORDER BY month",
            (branch_id,),
        )
        return [r[0] for r in cur.fetchall()]


def _build_weeks(
    days: list[dict[str, Any]],
    hours: list[int],
    bench_hours: set[int],
) -> list[dict[str, Any]]:
    if not days:
        return []
    first_day = date.fromisoformat(days[0]["date"])
    last_day = date.fromisoformat(days[-1]["date"])
    current = week_start_monday(first_day)
    weeks: list[dict[str, Any]] = []
    while current <= last_day:
        week_end = current + timedelta(days=6)
        week_days = []
        week_vals = []
        for day in days:
            day_date = date.fromisoformat(day["date"])
            if day_date < current or day_date > week_end:
                continue
            week_days.append(day)
            for idx, hour in enumerate(hours):
                if hour in bench_hours:
                    week_vals.append(day["cells"][idx]["load_pct"])
        week_avg = round(sum(week_vals) / len(week_vals), 2) if week_vals else 0.0
        weeks.append(
            {
                "week_start": current.isoformat(),
                "week_end": week_end.isoformat(),
                "days": week_days,
                "week_avg": week_avg,
            }
        )
        current += timedelta(days=7)
    return weeks


def month_payload(branch_id: int, month: str) -> dict[str, Any]:
    with get_hist_conn() as conn:
        cur = conn.execute(
            """
            SELECT resource_type, date, dow, hour, load_pct
            FROM historical_loads
            WHERE branch_id = ? AND month = ?
            """,
            (branch_id, month),
        )
        rows = cur.fetchall()
        if not rows:
            raise KeyError("Нет данных за выбранный месяц")
        cur2 = conn.execute(
            """
            SELECT resource_type
            FROM historical_types
            WHERE branch_id = ? AND month = ?
            ORDER BY order_index
            """,
            (branch_id, month),
        )
        type_order = [r[0] for r in cur2.fetchall()]

    hours = sorted(h for h in {int(r["hour"]) for r in rows} if 8 <= h <= 23)
    bench_hours = {h for h in hours if 10 <= h <= 21}
    dates = sorted({r["date"] for r in rows})
    hour_index = {hour: idx for idx, hour in enumerate(hours)}
    date_index = {day: idx for idx, day in enumerate(dates)}

    types_map: dict[str, list[dict[str, Any]]] = {}
    for day in dates:
        dow = next((int(r["dow"]) for r in rows if r["date"] == day), 1)
        for type_name in type_order:
            types_map.setdefault(type_name, [])
            cells = [{"load_pct": 0.0, "busy_count": 0, "staff_total": 0} for _ in hours]
            types_map[type_name].append(
                {
                    "date": day,
                    "dow": dow,
                    "cells": cells,
                }
            )

    for r in rows:
        type_name = r["resource_type"]
        day_idx = date_index.get(r["date"])
        hour_idx = hour_index.get(int(r["hour"]))
        if day_idx is None or hour_idx is None:
            continue
        types_map.setdefault(type_name, [])
        if day_idx >= len(types_map[type_name]):
            continue
        types_map[type_name][day_idx]["cells"][hour_idx]["load_pct"] = float(r["load_pct"])

    types_out = []
    for type_name in type_order or sorted(types_map.keys()):
        days = types_map.get(type_name) or []
        all_vals = []
        for day in days:
            bench_vals = [
                day["cells"][i]["load_pct"]
                for i, h in enumerate(hours)
                if h in bench_hours
            ]
            all_vals.extend(bench_vals)
            day_avg = round(sum(bench_vals) / len(bench_vals), 2) if bench_vals else 0.0
            day["day_avg"] = day_avg
        month_avg = round(sum(all_vals) / len(all_vals), 2) if all_vals else 0.0
        weeks = _build_weeks(days, hours, bench_hours)
        types_out.append(
            {
                "name": type_name,
                "weeks": weeks,
                "month_avg": month_avg,
            }
        )

    return {
        "branch_id": branch_id,
        "month": month,
        "hours": hours,
        "types": types_out,
    }
