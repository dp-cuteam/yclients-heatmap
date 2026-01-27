from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sqlite3
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl


def _normalize(value: str) -> str:
    return " ".join(value.strip().lower().replace("ё", "е").split())


def _load_branch_map() -> Dict[str, str]:
    path = Path(__file__).resolve().parents[1] / "data" / "branch_mapping.json"
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    mapping: Dict[str, str] = {}
    for item in data or []:
        name = item.get("name")
        code = item.get("code")
        if name and code:
            mapping[_normalize(str(name))] = str(code)
    return mapping


def _connect_db() -> Tuple[str, object]:
    db_url = os.getenv("CUTEAM_DATABASE_URL") or os.getenv("DATABASE_URL") or ""
    if db_url.startswith("postgres"):
        try:
            import psycopg
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError("psycopg is required for Postgres import") from exc
        conn = psycopg.connect(db_url)
        return "postgres", conn
    db_path = Path(os.getenv("CUTEAM_DB_PATH", "./data/cuteam.db"))
    conn = sqlite3.connect(db_path)
    return "sqlite", conn


def _upsert_value(
    conn: object, kind: str, branch: str, date_iso: str, value: float
) -> None:
    if kind == "postgres":
        sql = (
            "INSERT INTO manual_sheet_daily (branch_code, metric_code, date, value, source, updated_at) "
            "VALUES (%s, 'coffee_checks', %s, %s, 'manual', NOW()) "
            "ON CONFLICT (branch_code, metric_code, date) DO UPDATE "
            "SET value=EXCLUDED.value, source=EXCLUDED.source, updated_at=EXCLUDED.updated_at"
        )
        params = (branch, date_iso, float(value))
    else:
        sql = (
            "INSERT INTO manual_sheet_daily (branch_code, metric_code, date, value, source, updated_at) "
            "VALUES (?, 'coffee_checks', ?, ?, 'manual', datetime('now')) "
            "ON CONFLICT(branch_code, metric_code, date) DO UPDATE "
            "SET value=excluded.value, source=excluded.source, updated_at=excluded.updated_at"
        )
        params = (branch, date_iso, float(value))
    cur = conn.cursor()
    cur.execute(sql, params)
    conn.commit()


def _parse_date(value: object) -> Optional[str]:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d.%m.%y", "%d/%m/%y"):
            try:
                return dt.datetime.strptime(raw, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_sheet(path: Path, sheet_name: str | None = None) -> int:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    col_branch = None
    col_date = None
    col_checks = None
    for idx, cell in enumerate(header, start=1):
        if not isinstance(cell, str):
            continue
        label = _normalize(cell)
        if label == "торговое предприятие":
            col_branch = idx
        elif label == "учетный день":
            col_date = idx
        elif label == "чеков всего":
            col_checks = idx

    if not (col_branch and col_date and col_checks):
        raise RuntimeError("Expected columns not found in header row.")

    branch_map = _load_branch_map()
    kind, conn = _connect_db()
    inserted = 0
    unknown_branches = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_branch = row[col_branch - 1]
        raw_date = row[col_date - 1]
        raw_checks = row[col_checks - 1]
        if not raw_branch or raw_checks is None:
            continue
        if not isinstance(raw_branch, str):
            continue
        branch_code = branch_map.get(_normalize(raw_branch))
        if not branch_code:
            unknown_branches.add(raw_branch)
            continue
        date_iso = _parse_date(raw_date)
        if not date_iso:
            continue
        try:
            value = float(raw_checks)
        except (TypeError, ValueError):
            continue
        _upsert_value(conn, kind, branch_code, date_iso, value)
        inserted += 1

    conn.close()
    wb.close()
    if unknown_branches:
        print("Unknown branches:", ", ".join(sorted(unknown_branches)))
    return inserted


def main() -> None:
    parser = argparse.ArgumentParser(description="Import coffee checks into manual_sheet_daily.")
    parser.add_argument("path", help="Path to Checks.xlsx")
    parser.add_argument("--sheet", dest="sheet", default=None, help="Sheet name (optional)")
    args = parser.parse_args()

    count = parse_sheet(Path(args.path), args.sheet)
    print(f"Inserted/updated rows: {count}")


if __name__ == "__main__":
    main()
