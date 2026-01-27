"""Sync manual data from Google Sheets (or local .xlsx) into SQLite/Postgres."""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import os
import re
import sqlite3
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:  # noqa: BLE001
    psycopg = None
    dict_row = None

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from shared import reference  # noqa: E402


DATE_SERIAL_THRESHOLD = 20000  # Google/Excel serials for modern dates are > 20000.
DATE_EPOCH = dt.date(1899, 12, 30)
ARTICLE_HEADER = "\u0441\u0442\u0430\u0442\u044c\u044f"
DEFAULT_SHEET_NAME = "\u0418\u0422\u041e\u0413\u041e-26"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync manual sheet data into SQLite/Postgres.")
    parser.add_argument("--db", default=str(ROOT / "data" / "cuteam.db"))
    parser.add_argument("--sheet-id", default=os.environ.get("SHEET_ID"))
    parser.add_argument("--sheet-name", default=os.environ.get("SHEET_NAME", DEFAULT_SHEET_NAME))
    parser.add_argument("--service-account-json", default=os.environ.get("GOOGLE_SA_JSON"))
    parser.add_argument("--service-account-b64", default=os.environ.get("GOOGLE_SA_JSON_B64"))
    parser.add_argument("--xlsx", help="Local .xlsx path for offline sync")
    parser.add_argument("--date-from", dest="date_from", help="YYYY-MM-DD")
    parser.add_argument("--date-to", dest="date_to", help="YYYY-MM-DD")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def _is_postgres_target(value: str) -> bool:
    return value.startswith("postgres")


class DBConn:
    def __init__(self, conn, kind: str):
        self._conn = conn
        self._kind = kind

    def _prepare(self, sql: str) -> str:
        if self._kind == "postgres":
            return sql.replace("?", "%s")
        return sql

    def execute(self, sql: str, params: Iterable | None = None):
        params = [] if params is None else params
        sql = self._prepare(sql)
        if self._kind == "postgres":
            cur = self._conn.cursor()
            cur.execute(sql, params)
            return cur
        return self._conn.execute(sql, params)

    def executemany(self, sql: str, seq: Iterable[Iterable]):
        sql = self._prepare(sql)
        if self._kind == "postgres":
            cur = self._conn.cursor()
            cur.executemany(sql, seq)
            return cur
        return self._conn.executemany(sql, seq)

    def commit(self) -> None:
        self._conn.commit()

    def close(self) -> None:
        self._conn.close()


def _connect_db(target: str) -> DBConn:
    if _is_postgres_target(target):
        if not psycopg:
            raise RuntimeError("psycopg is required for Postgres sync")
        raw = psycopg.connect(target, row_factory=dict_row)
        return DBConn(raw, "postgres")
    db_path = Path(target)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    raw = sqlite3.connect(str(db_path))
    raw.row_factory = sqlite3.Row
    return DBConn(raw, "sqlite")


def load_google_values(sheet_id: str, sheet_name: str, key_path: Optional[str], key_b64: Optional[str]) -> List[List[Any]]:
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except Exception as exc:  # pragma: no cover - runtime import guard
        raise RuntimeError("Missing Google API dependencies. Install: pip install -r requirements.txt") from exc

    if not sheet_id:
        raise ValueError("sheet_id is required")

    if key_b64:
        info = json.loads(base64.b64decode(key_b64).decode("utf-8"))
        credentials = service_account.Credentials.from_service_account_info(
            info, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
        )
    elif key_path:
        credentials = service_account.Credentials.from_service_account_file(
            key_path, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
        )
    else:
        raise ValueError("Provide service account via --service-account-json or --service-account-b64")

    service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    result = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=sheet_id,
            range=sheet_name,
            valueRenderOption="UNFORMATTED_VALUE",
            dateTimeRenderOption="FORMATTED_STRING",
        )
        .execute()
    )
    values = result.get("values", [])
    return normalize_grid(values)


def load_xlsx_values(path: str, sheet_name: str) -> List[List[Any]]:
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover - runtime import guard
        raise RuntimeError("openpyxl is required for .xlsx import") from exc

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return normalize_grid(rows)


def normalize_grid(rows: List[List[Any]]) -> List[List[Any]]:
    if not rows:
        return []
    max_len = max(len(row) for row in rows)
    return [row + [None] * (max_len - len(row)) for row in rows]


def parse_date_cell(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        if value > DATE_SERIAL_THRESHOLD:
            return (DATE_EPOCH + dt.timedelta(days=int(round(value)))).isoformat()
        return None
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d.%m.%y", "%d/%m/%y"):
            try:
                return dt.datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(" ", "")
        if not s:
            return None
        s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


RU_MONTHS = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "мая": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def find_year_hint(header: List[Any]) -> Optional[int]:
    for cell in header:
        if isinstance(cell, (int, float)):
            year = int(cell)
            if 2000 <= year <= 2100:
                return year
        if isinstance(cell, str):
            s = cell.strip()
            if len(s) == 4 and s.isdigit():
                year = int(s)
                if 2000 <= year <= 2100:
                    return year
    return None


def parse_ru_header_date(value: Any, year_hint: Optional[int]) -> Optional[str]:
    if not year_hint:
        return None
    if not isinstance(value, str):
        return None
    s = value.strip().lower()
    if not s:
        return None
    match = re.search(r"(\d{1,2})\s+([а-яё\.]+)", s)
    if not match:
        return None
    day = int(match.group(1))
    mon_raw = match.group(2).strip(".")
    month = None
    for key, number in RU_MONTHS.items():
        if mon_raw.startswith(key):
            month = number
            break
    if not month:
        return None
    try:
        return dt.date(year_hint, month, day).isoformat()
    except ValueError:
        return None


def parse_date_arg(value: Optional[str]) -> Optional[dt.date]:
    if not value:
        return None
    return dt.date.fromisoformat(value)



def find_header_row(rows: List[List[Any]]) -> int:
    for idx, row in enumerate(rows[:12]):
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower() in {ARTICLE_HEADER, "article"}:
                return idx
    return 3  # fallback for current sheet layout


def find_column(row: List[Any], name: str) -> Optional[int]:
    target = name.strip().lower()
    for i, cell in enumerate(row):
        if isinstance(cell, str) and cell.strip().lower() == target:
            return i
    return None


def is_branch_header(
    row: List[Any],
    col_article: Optional[int],
    col_metric: Optional[int],
    known_branches: set[str],
    ignore_branches: set[str],
) -> Optional[str]:
    if not row:
        return None
    if col_article is not None and row[col_article] not in (None, ""):
        return None
    if col_metric is not None and row[col_metric] not in (None, ""):
        return None
    non_empty = [cell for cell in row if cell not in (None, "")]
    if not non_empty:
        return None
    for cell in non_empty:
        if isinstance(cell, str):
            candidate = cell.strip()
            if candidate in known_branches or candidate in ignore_branches:
                return candidate
    first = row[0]
    if isinstance(first, str):
        return first.strip() or None
    return None


def iter_records(
    rows: List[List[Any]],
    date_from: Optional[dt.date] = None,
    date_to: Optional[dt.date] = None,
) -> Tuple[List[Tuple[str, str, str, float, str]], Dict[str, int]]:
    header_idx = find_header_row(rows)
    header = rows[header_idx]

    col_metric = find_column(header, "metric_code")
    col_source = find_column(header, "source")
    col_article = find_column(header, ARTICLE_HEADER)
    col_branch = find_column(header, "branch_code")
    year_hint = find_year_hint(header)

    date_cols: List[Tuple[int, str]] = []
    for idx, cell in enumerate(header):
        parsed = parse_date_cell(cell)
        if not parsed:
            parsed = parse_ru_header_date(cell, year_hint)
        if parsed:
            date_cols.append((idx, parsed))

    if not date_cols:
        raise RuntimeError("No date columns detected. Check sheet headers.")

    if date_from or date_to:
        filtered: List[Tuple[int, str]] = []
        for idx, date_iso in date_cols:
            d = dt.date.fromisoformat(date_iso)
            if date_from and d < date_from:
                continue
            if date_to and d > date_to:
                continue
            filtered.append((idx, date_iso))
        date_cols = filtered

    known_branches = {b["code"] for b in reference.BRANCHES}
    ignore_branches = set(reference.IGNORE_BRANCH_CODES)
    ignore_labels = set(reference.IGNORE_LABELS)

    current_branch: Optional[str] = None
    skip_branch = False
    records: List[Tuple[str, str, str, float, str]] = []
    stats = {"rows": 0, "unknown_labels": 0, "unknown_branches": 0, "skipped_empty": 0}

    for row in rows[header_idx + 1 :]:
        stats["rows"] += 1
        if not any(cell not in (None, "") for cell in row):
            stats["skipped_empty"] += 1
            continue

        header_branch = is_branch_header(row, col_article, col_metric, known_branches, ignore_branches)
        if header_branch:
            current_branch = header_branch
            skip_branch = current_branch in ignore_branches
            if current_branch not in known_branches and current_branch not in ignore_branches:
                stats["unknown_branches"] += 1
            continue

        branch_code = None
        if col_branch is not None:
            val = row[col_branch]
            if val not in (None, ""):
                branch_code = str(val).strip()
        if not branch_code:
            branch_code = current_branch
        if not branch_code or branch_code in ignore_branches:
            continue
        if skip_branch:
            continue

        metric_code = None
        if col_metric is not None:
            raw = row[col_metric]
            if raw not in (None, ""):
                metric_code = str(raw).strip()

        label = None
        if col_article is not None:
            raw_label = row[col_article]
            if raw_label not in (None, ""):
                label = str(raw_label).strip()

        if label in ignore_labels:
            continue

        if not metric_code:
            if label:
                metric_code = reference.RAW_LABEL_TO_METRIC_CODE.get(label)
            if not metric_code:
                stats["unknown_labels"] += 1
                continue

        source = "manual"
        if col_source is not None:
            raw_source = row[col_source]
            if raw_source not in (None, ""):
                source = str(raw_source).strip().lower()
        if source != "manual":
            continue

        for col_idx, date_iso in date_cols:
            val = parse_number(row[col_idx] if col_idx < len(row) else None)
            if val is None:
                continue
            records.append((branch_code, metric_code, date_iso, val, source))

    stats["records"] = len(records)
    stats["date_cols"] = len(date_cols)
    return records, stats


def _split_sql_statements(sql: str) -> List[str]:
    # Drop line comments so ";" inside comments doesn't break splitting.
    lines = []
    for line in sql.splitlines():
        if line.lstrip().startswith("--"):
            continue
        lines.append(line)
    cleaned = "\n".join(lines)
    parts = [chunk.strip() for chunk in cleaned.split(";")]
    return [part for part in parts if part]


def ensure_schema(conn: DBConn) -> None:
    schema_path = ROOT / "shared" / "db" / "schema.sql"
    schema = schema_path.read_text(encoding="utf-8")
    if conn._kind == "postgres":
        schema = schema.replace("CREATE VIEW IF NOT EXISTS", "CREATE OR REPLACE VIEW")
    for stmt in _split_sql_statements(schema):
        conn.execute(stmt)
    conn.commit()


def upsert_records(conn: DBConn, records: Sequence[Tuple[str, str, str, float, str]]) -> int:
    if not records:
        return 0
    now = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
    rows = [(b, m, d, v, s, now) for (b, m, d, v, s) in records]
    sql = (
        "INSERT INTO manual_sheet_daily (branch_code, metric_code, date, value, source, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?) "
        "ON CONFLICT(branch_code, metric_code, date) DO UPDATE SET "
        "value=excluded.value, source=excluded.source, updated_at=excluded.updated_at"
    )
    conn.executemany(sql, rows)
    conn.commit()
    return len(rows)


def seed_dimensions(conn: DBConn) -> None:
    # Branches
    branch_sql = (
        "INSERT INTO branches (code, name) VALUES (?, ?) "
        "ON CONFLICT(code) DO UPDATE SET name=excluded.name"
    )
    conn.executemany(branch_sql, [(b["code"], b["name"]) for b in reference.BRANCHES])

    # Metrics
    metric_rows = []
    for label, code in reference.RAW_LABEL_TO_METRIC_CODE.items():
        metric_rows.append((code, label))
    for code, meta in reference.EXTRA_INPUT_METRICS.items():
        metric_rows.append((code, meta["label"]))
    for code, meta in reference.DERIVED_METRICS.items():
        metric_rows.append((code, meta["label"]))

    metric_sql = (
        "INSERT INTO metrics (code, label) VALUES (?, ?) ON CONFLICT(code) DO NOTHING"
        if conn._kind == "postgres"
        else "INSERT OR IGNORE INTO metrics (code, label) VALUES (?, ?)"
    )
    conn.executemany(metric_sql, metric_rows)
    conn.commit()


def main() -> None:
    args = parse_args()

    if args.xlsx:
        rows = load_xlsx_values(args.xlsx, args.sheet_name)
    else:
        rows = load_google_values(args.sheet_id, args.sheet_name, args.service_account_json, args.service_account_b64)

    date_from = parse_date_arg(args.date_from)
    date_to = parse_date_arg(args.date_to)
    records, stats = iter_records(rows, date_from=date_from, date_to=date_to)

    print(
        "Parsed rows: {rows}, date columns: {date_cols}, records: {records}, "
        "unknown labels: {unknown_labels}, unknown branches: {unknown_branches}".format(**stats)
    )

    if args.dry_run:
        return

    conn = _connect_db(args.db)
    try:
        ensure_schema(conn)
        seed_dimensions(conn)
        inserted = upsert_records(conn, records)
        print(f"Upserted {inserted} rows into manual_sheet_daily")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
