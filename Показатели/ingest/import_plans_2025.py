from __future__ import annotations

import argparse
import os
import sqlite3
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl


MONTHS = {
    "январ": 1,
    "феврал": 2,
    "март": 3,
    "апрел": 4,
    "ма": 5,  # май/мая
    "июн": 6,
    "июл": 7,
    "август": 8,
    "сентябр": 9,
    "октябр": 10,
    "ноябр": 11,
    "декабр": 12,
}


METRIC_LABELS = {
    "Выручка всего": "revenue_total",
    "Аренда": "revenue_open_space",
    "Кабинеты": "revenue_cabinets",
    "Лекторий": "revenue_lecture",
    "Лаборатория": "revenue_lab",
    "Ритейл": "revenue_retail",
    "Услуги салона": "revenue_salon",
    "Кофейня": "coffee_revenue_total",
}


def _normalize(value: str) -> str:
    return " ".join(value.strip().lower().replace("ё", "е").split())


def _parse_month(cell_value: Optional[str]) -> Optional[int]:
    if not isinstance(cell_value, str):
        return None
    raw = _normalize(cell_value)
    for key, number in MONTHS.items():
        if raw.startswith(key) or f" {key}" in raw:
            return number
    return None


def _load_branch_map() -> Dict[str, str]:
    path = Path(__file__).resolve().parents[1] / "data" / "branch_mapping.json"
    if not path.exists():
        return {}
    import json

    data = json.loads(path.read_text(encoding="utf-8"))
    mapping: Dict[str, str] = {}
    for item in data or []:
        name = item.get("name")
        code = item.get("code")
        if name and code:
            mapping[_normalize(str(name))] = str(code)
    return mapping


def _connect_db() -> Tuple[str, object]:
    db_url = os.getenv("CUTEAM_DATABASE_URL") or os.getenv("DATABASE_URL") or ""
    if db_url.startswith("postgres"):
        try:
            import psycopg
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError("psycopg is required for Postgres plans import") from exc
        conn = psycopg.connect(db_url)
        return "postgres", conn
    db_path = Path(os.getenv("CUTEAM_DB_PATH", "./data/cuteam.db"))
    conn = sqlite3.connect(db_path)
    return "sqlite", conn


def _upsert_plan(conn: object, kind: str, branch: str, month: str, code: str, value: float) -> None:
    if kind == "postgres":
        sql = (
            "INSERT INTO plans_monthly (branch_code, metric_code, month_start, value, updated_at) "
            "VALUES (%s, %s, %s, %s, NOW()) "
            "ON CONFLICT (branch_code, metric_code, month_start) DO UPDATE "
            "SET value=EXCLUDED.value, updated_at=EXCLUDED.updated_at"
        )
    else:
        sql = (
            "INSERT INTO plans_monthly (branch_code, metric_code, month_start, value, updated_at) "
            "VALUES (?, ?, ?, ?, datetime('now')) "
            "ON CONFLICT(branch_code, metric_code, month_start) DO UPDATE "
            "SET value=excluded.value, updated_at=excluded.updated_at"
        )
    cur = conn.cursor()
    cur.execute(sql, (branch, code, f"{month}-01", float(value)))
    conn.commit()


def parse_sheet(path: Path, sheet_name: str | None = None) -> int:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    month_cols: Dict[int, int] = {}
    for idx, cell in enumerate(header, start=1):
        month = _parse_month(cell)
        if month:
            month_cols[month] = idx

    if not month_cols:
        raise RuntimeError("No month columns detected in header row.")

    branch_map = _load_branch_map()
    kind, conn = _connect_db()
    current_branch = None
    inserted = 0

    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if isinstance(label, str) and label.strip():
            label_norm = _normalize(label)
            if label_norm in branch_map:
                current_branch = branch_map[label_norm]
                continue

        if not current_branch:
            continue

        if not isinstance(label, str):
            continue
        metric_code = METRIC_LABELS.get(label.strip())
        if not metric_code:
            continue

        for month, col_idx in month_cols.items():
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                continue
            month_key = f"2025-{month:02d}"
            _upsert_plan(conn, kind, current_branch, month_key, metric_code, numeric)
            inserted += 1

    wb.close()
    conn.close()
    return inserted


def main() -> None:
    parser = argparse.ArgumentParser(description="Import 2025 plans from Excel into plans_monthly.")
    parser.add_argument("path", help="Path to plans25.xlsx")
    parser.add_argument("--sheet", dest="sheet", default=None, help="Sheet name (optional)")
    args = parser.parse_args()

    count = parse_sheet(Path(args.path), args.sheet)
    print(f"Inserted/updated plan cells: {count}")


if __name__ == "__main__":
    main()
